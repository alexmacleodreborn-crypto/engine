"""
A7DO Genesis Mind — v5 Interactive Dashboard
Fixed: sidebar/pages/tick structure
New: Camera, Mic, Speaker, Auto-step, Master Dashboard default
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import openpyxl
import math
import time
from pathlib import Path

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="A7DO Genesis Mind",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Workbook path ─────────────────────────────────────────────────────────────
XLSX = Path("excel_report/a7do-final/A7DO_DNA_Master_v5_FINAL.xlsx")

@st.cache_resource
def load_workbook():
    return openpyxl.load_workbook(XLSX, data_only=True)

@st.cache_data
def sheet_to_df(sheet_name):
    wb = load_workbook()
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            rows.append(list(row))
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)

@st.cache_data
def get_sheet_names():
    wb = load_workbook()
    return wb.sheetnames

# ── Organism state (pure Python, tick-driven) ─────────────────────────────────
def organism_state(tick):
    week = round(tick / 80)
    height = 50 if week < 40 else min(50 + (177-50)*(1-math.exp(-0.005*(week-40))), 177)
    mass   = 3.5 if week < 40 else min(3.5 + (70-3.5)*(1-math.exp(-0.004*(week-40))), 70)
    hr     = 140 if week < 40 else max(70, 140-(140-70)*((week-40)/1160))
    vocab  = round(50000/(1+math.exp(-0.05*(week-156)))) if week > 0 else 0
    motor  = 5 if week>=400 else (4 if week>=260 else (3 if week>=160 else (2 if week>=80 else 1)))
    tom    = 5 if week>=624 else (4 if week>=312 else (3 if week>=208 else (2 if week>=156 else 1)))
    perm   = 3 if week>=52  else (2 if week>=44  else (1 if week>=36  else 0))
    birth  = tick >= 3200
    phase7 = tick >= 96000
    wisdom = min(0.1+((week-1200)/800)*0.9, 1.0) if week >= 1200 else 0.0
    if   week >= 1200: stage = "Mature Adult"
    elif week >= 1100: stage = "Mid Adult"
    elif week >= 1000: stage = "Adult"
    elif week >= 936:  stage = "Young Adult"
    elif week >= 624:  stage = "Adolescent"
    elif week >= 260:  stage = "Pre-Adolescent"
    elif week >= 156:  stage = "Child"
    elif week >= 80:   stage = "Toddler"
    elif week >= 52:   stage = "Infant"
    elif week >= 40:   stage = "Newborn"
    elif week >= 28:   stage = "Fetal Late"
    elif week >= 12:   stage = "Fetal Mid"
    elif week >= 4:    stage = "Fetal Early"
    else:              stage = "Embryo"
    if tick % 800 == 0:  ll = "💤 Sleep Consolidation"
    elif tick % 10 == 0: ll = "🔁 Repetition"
    elif tick % 5 == 0:  ll = "🤝 Interaction"
    else:                ll = "👁️ Exposure"
    C        = min(0.05 + week/3000, 1.0)
    pred_err = max(0.1, math.exp(-0.0001*tick))
    ltm      = min(int(tick * 0.96), 200000)
    return dict(tick=tick, week=week, stage=stage, height=round(height,1),
                mass=round(mass,2), hr=round(hr,1), vocab=vocab,
                motor=motor, tom=tom, perm=perm, birth=birth,
                phase7=phase7, wisdom=round(wisdom,3), ll_phase=ll,
                C=round(C,3), pred_err=round(pred_err,3), ltm=ltm)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stSidebar"] { background: #0d1117; }
[data-testid="stSidebar"] * { color: #e6edf3 !important; }
.metric-card {
    background: linear-gradient(135deg, #1f2937, #111827);
    border: 1px solid #374151; border-radius: 12px;
    padding: 16px; text-align: center; margin: 4px;
}
.metric-val { font-size: 1.4rem; font-weight: 700; color: #60a5fa; }
.metric-lbl { font-size: 0.72rem; color: #9ca3af; margin-top: 4px; }
.phase-badge {
    display: inline-block; padding: 4px 10px; border-radius: 20px;
    font-size: 0.78rem; font-weight: 600; margin: 2px;
}
.ph-core    { background:#1e3a5f; color:#60a5fa; }
.ph-bio     { background:#1a3a2a; color:#4ade80; }
.ph-cog     { background:#3a1a3a; color:#c084fc; }
.ph-p3      { background:#3a2a1a; color:#fb923c; }
.ph-world   { background:#1a3a3a; color:#22d3ee; }
.ph-meta    { background:#2a2a1a; color:#facc15; }
.ph-loop    { background:#1a2a3a; color:#38bdf8; }
.ph-p7      { background:#2a1a1a; color:#f87171; }
.ph-p8      { background:#1a2a1a; color:#86efac; }
.section-hdr {
    font-size: 1.05rem; font-weight: 700;
    border-left: 4px solid #60a5fa;
    padding-left: 10px; margin: 14px 0 6px 0; color: #e2e8f0;
}
.auto-step-active { animation: pulse 1s infinite; }
@keyframes pulse { 0%,100%{opacity:1} 50%{opacity:0.6} }
</style>
""", unsafe_allow_html=True)

# ── Session state initialisation ──────────────────────────────────────────────
if "tick" not in st.session_state:
    st.session_state.tick = 10000
if "page" not in st.session_state:
    st.session_state.page = "🏠 Master Dashboard"
if "auto_step" not in st.session_state:
    st.session_state.auto_step = False
if "auto_speed" not in st.session_state:
    st.session_state.auto_speed = 80

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🧬 A7DO Genesis Mind")
    st.markdown("**v5 FINAL · 47 sheets · 337 formulas**")
    st.divider()

    # Tick slider
    tick = st.slider("⚡ Current Tick", 0, 160000,
                     st.session_state.tick, step=80,
                     help="Drag to advance A7DO through its lifecycle")
    st.session_state.tick = tick
    s = organism_state(tick)
    st.caption(f"Week {s['week']} · {s['stage']}")
    st.divider()

    # Navigation
    PAGES = [
        "🏠 Master Dashboard",
        "📈 Growth Timeline",
        "🧬 Biology",
        "🧠 Cognition & Phase 4",
        "🔄 Learning Loop",
        "🦿 Movement Engine",
        "🗣️ Word Learning Engine",
        "🔊 Speech Production",
        "🌐 Web-Hook Learning",
        "🌍 World & Social",
        "⚙️ Engines & Runtime",
        "✨ Phase 7 — Wisdom",
        "🚀 Phase 8 — AGI",
        "📷 Sensors (Camera/Mic)",
        "📊 All Sheets Explorer",
    ]
    page = st.radio("Navigate", PAGES,
                    index=PAGES.index(st.session_state.page),
                    label_visibility="collapsed")
    st.session_state.page = page
    st.divider()
    st.markdown("**47 sheets · 26,869 cells**")
    st.markdown("- Consolidated v17 (18)")
    st.markdown("- Master v3 (5)")
    st.markdown("- Consolidated v9 (5)")
    st.markdown("- Phase 7 (4) · Language (5)")
    st.markdown("- Phase 8 (7) · Sensors (1)")

state = organism_state(tick)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MASTER DASHBOARD (default)
# ═══════════════════════════════════════════════════════════════════════════════
if page == "🏠 Master Dashboard":
    st.title("🧬 A7DO Genesis Mind — Master Dashboard")
    st.caption(f"Tick {tick:,} · Week {state['week']} · {state['stage']} · {state['ll_phase']}")

    # Status banner
    if state['phase7']:
        st.success("🌟 PHASE 7 ACTIVE — Creative Synthesis · Wisdom · Career · Legacy engines online")
    elif state['birth']:
        st.info(f"🔄 Learning Loop: {state['ll_phase']}")
    else:
        st.warning("⏳ Prenatal — Birth at Tick 3,200")

    # ── Auto-step controls ──
    st.markdown('<div class="section-hdr">⚡ Auto-Step Control</div>', unsafe_allow_html=True)
    col_a, col_b, col_c, col_d = st.columns([1,1,2,3])
    with col_a:
        if st.button("▶ Start" if not st.session_state.auto_step else "⏸ Pause",
                     use_container_width=True):
            st.session_state.auto_step = not st.session_state.auto_step
    with col_b:
        if st.button("⏹ Reset", use_container_width=True):
            st.session_state.tick = 0
            st.session_state.auto_step = False
            st.rerun()
    with col_c:
        speed = st.select_slider("Step size", options=[80,400,800,1600,4000,8000],
                                  value=st.session_state.auto_speed, label_visibility="collapsed")
        st.session_state.auto_speed = speed
    with col_d:
        st.caption(f"Step: {speed} ticks ({speed//80} weeks) per advance")

    # Auto-step execution
    if st.session_state.auto_step:
        new_tick = min(st.session_state.tick + st.session_state.auto_speed, 160000)
        if new_tick >= 160000:
            st.session_state.auto_step = False
        st.session_state.tick = new_tick
        time.sleep(0.3)
        st.rerun()

    st.divider()

    # ── Key metrics ──
    metrics = [
        ("📏 Height",    f"{state['height']} cm"),
        ("⚖️ Mass",      f"{state['mass']} kg"),
        ("❤️ Heart Rate",f"{state['hr']} bpm"),
        ("💬 Vocab",     f"{state['vocab']:,}"),
        ("🦾 Motor",     f"{state['motor']}/5"),
        ("🧠 ToM",       f"{state['tom']}/5"),
        ("✨ Conscious", f"{state['C']}"),
        ("🦉 Wisdom",    f"{state['wisdom']}"),
        ("💾 LTM",       f"{state['ltm']:,}"),
        ("🔮 Pred Err",  f"{state['pred_err']}"),
    ]
    cols = st.columns(5)
    for i, (lbl, val) in enumerate(metrics):
        cols[i%5].markdown(f"""<div class="metric-card">
            <div class="metric-val">{val}</div>
            <div class="metric-lbl">{lbl}</div>
        </div>""", unsafe_allow_html=True)

    st.divider()

    # ── Growth curves ──
    col1, col2 = st.columns(2)
    ticks_r = list(range(0, 160001, 800))
    weeks_r = [t/80 for t in ticks_r]

    with col1:
        st.markdown('<div class="section-hdr">📏 Physical Growth</div>', unsafe_allow_html=True)
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=weeks_r,
            y=[organism_state(t)['height'] for t in ticks_r],
            name="Height (cm)", line=dict(color="#60a5fa",width=2)))
        fig.add_trace(go.Scatter(x=weeks_r,
            y=[organism_state(t)['mass'] for t in ticks_r],
            name="Mass (kg)", line=dict(color="#4ade80",width=2), yaxis="y2"))
        fig.add_vline(x=state['week'], line_dash="dash", line_color="#f87171")
        fig.update_layout(template="plotly_dark", height=260,
            yaxis=dict(title="Height cm"),
            yaxis2=dict(title="Mass kg", overlaying="y", side="right"),
            legend=dict(orientation="h",y=1.1), margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-hdr">💬 Vocabulary Growth</div>', unsafe_allow_html=True)
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=weeks_r,
            y=[organism_state(t)['vocab'] for t in ticks_r],
            fill="tozeroy", line=dict(color="#c084fc",width=2), name="V(t)"))
        fig2.add_vline(x=state['week'], line_dash="dash", line_color="#f87171")
        fig2.update_layout(template="plotly_dark", height=260,
            yaxis_title="Words", margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig2, use_container_width=True)

    # ── Phase timeline ──
    st.markdown('<div class="section-hdr">🗺️ Developmental Phase Timeline</div>', unsafe_allow_html=True)
    phases = [
        ("Phase 1 — Biology",       0,     3200,  "#1e40af"),
        ("Phase 2 — Sensorimotor",  3200,  6400,  "#065f46"),
        ("Phase 3 — Core Cognition",6400,  12480, "#7c2d12"),
        ("Phase 4 — Social Cog",    12480, 49920, "#581c87"),
        ("Phase 5 — Cultural",      49920, 74880, "#713f12"),
        ("Phase 6 — Identity",      74880, 96000, "#134e4a"),
        ("Phase 7 — Wisdom",        96000, 160000,"#7f1d1d"),
    ]
    fig3 = go.Figure()
    for name, start, end, color in phases:
        fig3.add_trace(go.Bar(x=[end-start], y=[name], base=[start],
            orientation='h', marker_color=color, name=name,
            hovertemplate=f"{name}<br>Tick {start:,}–{end:,}<extra></extra>"))
    fig3.add_vline(x=tick, line_color="#f87171", line_width=3,
                   annotation_text=f"Tick {tick:,}")
    fig3.update_layout(template="plotly_dark", height=260, showlegend=False,
                       barmode="overlay", xaxis_title="Tick",
                       margin=dict(l=0,r=0,t=20,b=0))
    st.plotly_chart(fig3, use_container_width=True)

    # ── Architecture completeness ──
    st.markdown('<div class="section-hdr">📊 Architecture Completeness</div>', unsafe_allow_html=True)
    domains = ["Biological body","Neural dynamics","Language & grounding","Cognitive architecture",
               "Energy & metabolism","Memory systems","Perception & vision","Motor intelligence",
               "Consciousness & self","Social & ToM","World & continuity","Temporal cognition"]
    v1  = [85,80,65,60,60,55,55,50,50,30,25,20]
    ph8 = [88,83,78,80,62,65,82,78,72,68,70,45]
    fig4 = go.Figure()
    fig4.add_trace(go.Bar(name="v1.0", x=v1, y=domains, orientation="h",
                           marker_color="#374151", text=[f"{s}%" for s in v1], textposition="inside"))
    fig4.add_trace(go.Bar(name="Phase 8", x=ph8, y=domains, orientation="h",
                           marker_color="#4ade80", text=[f"{s}%" for s in ph8], textposition="outside"))
    fig4.update_layout(template="plotly_dark", height=380, barmode="overlay",
                       xaxis=dict(range=[0,100],title="Completeness %"),
                       margin=dict(l=0,r=0,t=20,b=0))
    st.plotly_chart(fig4, use_container_width=True)

    # ── Sheet index ──
    st.markdown('<div class="section-hdr">🗂️ All 47 Sheets</div>', unsafe_allow_html=True)
    sheet_cats = {
        "🏠 Master Dashboard":"Core","🧬 DNA Loop Engine":"Core","📈 Growth Timeline":"Core",
        "🧮 Mathematical Framework":"Core","⚙️ Parameters":"Core","📖 System Specification":"Core",
        "🧬 Prenatal & Genesis":"Biology","📐 Spatial Coordinates":"Biology","🫀 Body Systems":"Biology",
        "🔬 Subsystems":"Biology","⚡ Energy & Metabolism":"Biology","📏 Full Growth System":"Biology",
        "🧠 Mind & Cognition":"Cognition","🗣️ Language Grounding":"Phase 4","🧠 Theory of Mind":"Phase 4",
        "🔮 Predictive Simulation":"Phase 4","🗺️ Scene Graph & Places":"Phase 4",
        "👁️ Object Permanence":"Phase 3","🦴 Proprioception (P1)":"Phase 3",
        "💾 Episodic Memory (P3)":"Phase 3","🎯 Value System (P4)":"Phase 3","🦾 Motor Planning (P6)":"Phase 3",
        "🦿 Movement Engine":"Phase 3",
        "🌍 World Systems":"World","🗺️ World Data":"World","📐 World & Space":"World",
        "🏥 Immersive Places":"World","👥 NPC Engine":"Social",
        "🔗 System Connections":"Meta","🔗 System Integration":"Meta",
        "🚀 v2.0 Architecture Audit":"Meta","⚙️ Runtime Patch v0.1":"Runtime",
        "🔄 Learning Loop":"Learning Loop",
        "🗣️ Word Learning Engine":"Learning Loop","🔊 Speech Production Engine":"Learning Loop",
        "🌐 Web-Hook Learning Pipeline":"Learning Loop",
        "👁️ Vision System (V1)":"Phase 8","🦾 Motor Intelligence (M1)":"Phase 8",
        "🗺️ Planning Engine (P2)":"Phase 8","🤝 Social Cognition (S2)":"Phase 8",
        "🪞 Meta-Cognition (R2)":"Phase 8","🌐 World Model Integration (W2)":"Phase 8",
        "📊 Phase 8 AGI Readiness":"Phase 8",
        "✨ Creative Synthesis Engine":"Phase 7","🦉 Wisdom Index Engine":"Phase 7",
        "🎯 Career Specialisation Engine":"Phase 7","🌟 Legacy Projection Engine":"Phase 7",
    }
    cat_cls = {"Core":"ph-core","Biology":"ph-bio","Cognition":"ph-cog","Phase 4":"ph-cog",
               "Phase 3":"ph-p3","World":"ph-world","Social":"ph-world","Meta":"ph-meta",
               "Runtime":"ph-meta","Learning Loop":"ph-loop","Phase 7":"ph-p7","Phase 8":"ph-p8"}
    badges = "".join(f'<span class="phase-badge {cat_cls.get(cat,"ph-core")}">{s}</span>'
                     for s,cat in sheet_cats.items())
    st.markdown(badges, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: GROWTH TIMELINE
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📈 Growth Timeline":
    st.title("📈 Growth Timeline")
    col1, col2 = st.columns(2)
    ticks_r = list(range(0, 160001, 400))
    weeks_r = [t/80 for t in ticks_r]
    with col1:
        st.subheader("Physical Development")
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=weeks_r, y=[organism_state(t)['height'] for t in ticks_r],
            name="Height cm", line=dict(color="#60a5fa",width=2)))
        fig.add_trace(go.Scatter(x=weeks_r, y=[organism_state(t)['mass'] for t in ticks_r],
            name="Mass kg", line=dict(color="#4ade80",width=2), yaxis="y2"))
        fig.add_trace(go.Scatter(x=weeks_r, y=[organism_state(t)['hr'] for t in ticks_r],
            name="HR bpm", line=dict(color="#f87171",width=2), yaxis="y3"))
        fig.add_vline(x=state['week'], line_dash="dash", line_color="white")
        fig.update_layout(template="plotly_dark", height=320,
            yaxis=dict(title="Height cm"),
            yaxis2=dict(title="Mass kg", overlaying="y", side="right"),
            yaxis3=dict(title="HR bpm", overlaying="y", side="right", position=0.95),
            legend=dict(orientation="h",y=1.1), margin=dict(l=0,r=0,t=30,b=0))
        st.plotly_chart(fig, use_container_width=True)
    with col2:
        st.subheader("Cognitive Development")
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=weeks_r, y=[organism_state(t)['vocab'] for t in ticks_r],
            fill="tozeroy", line=dict(color="#c084fc",width=2), name="Vocabulary"))
        fig2.add_trace(go.Scatter(x=weeks_r, y=[organism_state(t)['motor']*10000 for t in ticks_r],
            name="Motor ×10k", line=dict(color="#fb923c",width=2,dash="dot")))
        fig2.add_trace(go.Scatter(x=weeks_r, y=[organism_state(t)['tom']*10000 for t in ticks_r],
            name="ToM ×10k", line=dict(color="#22d3ee",width=2,dash="dot")))
        fig2.add_vline(x=state['week'], line_dash="dash", line_color="white")
        fig2.update_layout(template="plotly_dark", height=320,
            legend=dict(orientation="h",y=1.1), margin=dict(l=0,r=0,t=30,b=0))
        st.plotly_chart(fig2, use_container_width=True)

    st.subheader("Key Lifecycle Milestones")
    milestones = [
        (0,0,"Embryo","Fertilisation"),
        (320,4,"Embryo","Neural plate formed"),
        (640,8,"Embryo","Heart chamber partition"),
        (960,12,"Fetal Early","Limb differentiation"),
        (1600,20,"Fetal Mid","Sensory organs activating"),
        (2240,28,"Fetal Late","Neural gain rising"),
        (2880,36,"Fetal Late","Reflex arcs active"),
        (3200,40,"Newborn","Birth — umbilical severed"),
        (4160,52,"Infant","Object permanence partial"),
        (6400,80,"Toddler","Crawling → standing"),
        (10000,125,"Toddler","Language accelerating"),
        (12480,156,"Child","Desire ToM onset"),
        (16640,208,"Child","Belief ToM onset"),
        (24960,312,"Pre-Adolescent","Second-order ToM onset"),
        (32000,400,"Pre-Adolescent","Skilled motor stage 5"),
        (49920,624,"Adolescent","Full ToM online"),
        (74880,936,"Young Adult","All adolescent systems 100%"),
        (96000,1200,"Mature Adult","Phase 7 ACTIVATES"),
        (160000,2000,"Mature Adult","W(t)=0.95 — Legacy L(t)=0.94"),
    ]
    df_m = pd.DataFrame(milestones, columns=["Tick","Week","Stage","Milestone"])
    df_m["Status"] = df_m["Tick"].apply(lambda t: "✅" if t <= tick else "⏳")
    st.dataframe(df_m, use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: BIOLOGY
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🧬 Biology":
    st.title("🧬 Biology — Body Systems & Growth")
    tab1,tab2,tab3,tab4,tab5 = st.tabs(["🫀 Body Systems","🧬 Prenatal","⚡ Energy","📏 Full Growth","🔬 Subsystems"])
    with tab1:
        st.dataframe(sheet_to_df("🫀 Body Systems"), use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(sheet_to_df("🧬 Prenatal & Genesis"), use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(sheet_to_df("⚡ Energy & Metabolism"), use_container_width=True, hide_index=True)
    with tab4:
        st.dataframe(sheet_to_df("📏 Full Growth System"), use_container_width=True, hide_index=True)
    with tab5:
        st.dataframe(sheet_to_df("🔬 Subsystems"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: COGNITION & PHASE 4
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🧠 Cognition & Phase 4":
    st.title("🧠 Cognition & Phase 4")
    tab1,tab2,tab3,tab4,tab5,tab6,tab7 = st.tabs([
        "🧠 Mind","🗣️ Language","🧠 ToM","🔮 Predictive","🗺️ Scene Graph","👁️ Object Perm","💾 Episodic"])
    with tab1:
        col1,col2,col3,col4 = st.columns(4)
        col1.metric("Vocabulary",f"{state['vocab']:,}")
        col2.metric("Consciousness",state['C'])
        col3.metric("Pred Error",state['pred_err'])
        col4.metric("LTM Events",f"{state['ltm']:,}")
        st.dataframe(sheet_to_df("🧠 Mind & Cognition"), use_container_width=True, hide_index=True)
    with tab2:
        weeks_r = list(range(0,1300,10))
        vocabs  = [round(50000/(1+math.exp(-0.05*(w-156)))) for w in weeks_r]
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=weeks_r,y=vocabs,fill="tozeroy",line=dict(color="#c084fc",width=2)))
        fig.add_vline(x=state['week'],line_dash="dash",line_color="#f87171",
                      annotation_text=f"Wk {state['week']}: {state['vocab']:,}")
        fig.update_layout(template="plotly_dark",height=220,margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🗣️ Language Grounding"), use_container_width=True, hide_index=True)
    with tab3:
        tom_labels = {1:"Proto-ToM",2:"Desire inference",3:"False belief",4:"Second-order",5:"Full social"}
        st.success(f"ToM Stage {state['tom']} — {tom_labels[state['tom']]}")
        st.dataframe(sheet_to_df("🧠 Theory of Mind"), use_container_width=True, hide_index=True)
    with tab4:
        ticks_r = list(range(0,160001,800))
        errors  = [max(0.1,math.exp(-0.0001*t)) for t in ticks_r]
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=ticks_r,y=errors,fill="tozeroy",line=dict(color="#f87171",width=2)))
        fig.add_vline(x=tick,line_dash="dash",line_color="white")
        fig.update_layout(template="plotly_dark",height=220,margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🔮 Predictive Simulation"), use_container_width=True, hide_index=True)
    with tab5:
        st.dataframe(sheet_to_df("🗺️ Scene Graph & Places"), use_container_width=True, hide_index=True)
    with tab6:
        st.metric("Object Permanence Stage",f"{state['perm']}/3")
        st.dataframe(sheet_to_df("👁️ Object Permanence"), use_container_width=True, hide_index=True)
    with tab7:
        st.metric("LTM Events",f"{state['ltm']:,}")
        st.dataframe(sheet_to_df("💾 Episodic Memory (P3)"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: LEARNING LOOP
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔄 Learning Loop":
    st.title("🔄 Learning Loop — Experience-First Architecture")
    phase_colors = {"👁️ Exposure":"#1e3a5f","🤝 Interaction":"#065f46",
                    "🔁 Repetition":"#581c87","💤 Sleep Consolidation":"#1a1a2e"}
    cols = st.columns(4)
    for col,(ph,color) in zip(cols,phase_colors.items()):
        active = ph == state['ll_phase']
        col.markdown(f"""<div style="background:{color};border:{'3px solid #f87171' if active else '1px solid #374151'};
            border-radius:12px;padding:14px;text-align:center;">
            <div style="font-size:1.3rem">{ph.split()[0]}</div>
            <div style="font-weight:700;color:{'#f87171' if active else '#e2e8f0'}">{ph}</div>
            <div style="font-size:0.7rem;color:#9ca3af">{'← ACTIVE' if active else ''}</div>
        </div>""", unsafe_allow_html=True)
    st.divider()
    tab1,tab2,tab3 = st.tabs(["4-Stage Cycle","Word Pipeline","Identity"])
    with tab1:
        stages = {"Stage":["1 Exposure","2 Interaction","3 Repetition","4 Sleep"],
                  "Fires":["Every tick","Every 5t","Every 10t","Every 800t"],
                  "Engines":["EQ_SENS_06·EQ_EMOT_07·EQ_ATTN_09","EQ_PRED_08·Motor·Value TD(λ)",
                              "EQ_LANG_12·G(word)·Scene Graph·ToM","Episodic·EQ_PRED_08 dream replay"],
                  "Output":["Sensory vector·attention·emotional delta","Reinforcement·motor update",
                             "Word-object bindings·concept formation","Stable knowledge·grounded language"]}
        st.dataframe(pd.DataFrame(stages), use_container_width=True, hide_index=True)
    with tab2:
        steps = {"Step":["1 Exposure","2 Attention","3 Prediction","4 Emotion","5 Episodic","6 Repetition","7 Sleep","8 Grounded"],
                 "Process":["Hears 'ball' + sees object","Attention spikes on novel","Prediction error spikes",
                             "Emotional valence tags event","Stored in episodic memory","Multiple exposures strengthen",
                             "Dream replay binds word→percept","Stable word-object binding"],
                 "Engine":["EQ_SENS_06","EQ_ATTN_09","EQ_PRED_08","EQ_EMOT_07","Episodic","EQ_LANG_12","EQ_PRED_08","G(word)"],
                 "Status":["✅","✅","✅","✅","✅","🟡","✅","🟡"]}
        st.dataframe(pd.DataFrame(steps), use_container_width=True, hide_index=True)
    with tab3:
        identity = {"Component":["Episodic self-history","Stable prediction","Emotional history",
                                  "Social bonds","Value system","Self-model","Skill identity","Wisdom"],
                    "Status":[f"✅ {state['ltm']:,} events",f"✅ Error={state['pred_err']}",
                               "✅ H(t) active","✅ Lorraine=0.95","✅ V(s)=1.0",
                               f"✅ C={state['C']}",f"🟡 Motor {state['motor']}",
                               "✅ Active" if state['phase7'] else "⏳ Wk 1200+"]}
        st.dataframe(pd.DataFrame(identity), use_container_width=True, hide_index=True)
    st.subheader("Full Sheet")
    st.dataframe(sheet_to_df("🔄 Learning Loop"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: MOVEMENT ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🦿 Movement Engine":
    st.title("🦿 Movement Engine — Sections 4.1–4.10")
    motor_labels = {1:"Reflex-only",2:"Crawling",3:"Walking",4:"Coordinated",5:"Skilled"}
    st.info(f"Motor Stage {state['motor']} — {motor_labels[state['motor']]}")
    tab1,tab2,tab3,tab4 = st.tabs(["4.1–4.3 Dynamics","4.4 Control Stack","4.5–4.6 Primitives & Reflexes","4.7–4.8 FOV & Learning"])
    with tab1:
        st.code("""M(θ)θ̈ + C(θ,θ̇) + G(θ) + Jc(θ)ᵀλ = τ(θ,θ̇,A)
F_m = A_m · F_max,m · f_length(l_m) · f_velocity(v_m)
τ(θ,θ̇,A) = Σ_m B_m(θ) · F_m""", language="text")
        l_range = [i/100 for i in range(50,151)]
        f_len = [math.exp(-(l-1.0)**2/(2*0.04)) for l in l_range]
        fig = go.Figure(go.Scatter(x=l_range,y=f_len,fill="tozeroy",line=dict(color="#4ade80",width=2)))
        fig.update_layout(template="plotly_dark",height=200,xaxis_title="Muscle Length",
                          yaxis_title="Force Scale",margin=dict(l=0,r=0,t=10,b=0))
        st.plotly_chart(fig,use_container_width=True)
    with tab2:
        layers = {"Layer":["1 Intent","2 Task Error","3 Task PD","4 IK","5 Inv Dynamics","6 Torque→Muscle"],
                  "Formula":["Goal → x_target","e_x = x_target − x(θ)","ẍ_des = Kp·e_x + Kd·ė_x",
                              "θ̈_des = J†·ẍ_des + (I−J†J)·θ̈_null",
                              "τ_des = M·θ̈_des + C + G + τ_stab",
                              "min ||τ_des − τ(A)||² + λ||A||²"]}
        st.dataframe(pd.DataFrame(layers), use_container_width=True, hide_index=True)
    with tab3:
        prims = {"Primitive":["Posture","Balance","Gait","Reaching","Grasp"],
                 "Active":["✅" if state['motor']>=1 else "🔒","✅" if state['motor']>=2 else "🔒",
                            "✅" if state['motor']>=3 else "🔒","✅" if state['motor']>=2 else "🔒",
                            "✅" if state['motor']>=4 else "🔒"],
                 "Stage":[1,2,3,2,4]}
        st.dataframe(pd.DataFrame(prims), use_container_width=True, hide_index=True)
    with tab4:
        st.code("""e_gaze = x_fovea,target − x_fovea,current
θ̈_head = Kp_gaze·e_gaze − Kd_gaze·θ̇_head
r_t = w₁(−||e_x||) + w₂(−energy) + w₃(balance) + w₄(−FOV_err)""", language="text")
    st.subheader("Full Sheet")
    st.dataframe(sheet_to_df("🦿 Movement Engine"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: WORD LEARNING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🗣️ Word Learning Engine":
    st.title("🗣️ Word Learning Engine")
    col1,col2 = st.columns(2)
    with col1:
        weeks_r = list(range(0,1300,5))
        vocabs  = [round(50000/(1+math.exp(-0.05*(w-156)))) for w in weeks_r]
        fig = go.Figure(go.Scatter(x=weeks_r,y=vocabs,fill="tozeroy",line=dict(color="#4ade80",width=2)))
        fig.add_vline(x=state['week'],line_dash="dash",line_color="#f87171",
                      annotation_text=f"{state['vocab']:,} words")
        fig.update_layout(template="plotly_dark",height=250,xaxis_title="Week",
                          yaxis_title="Words",margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
    with col2:
        conf_vals = [0.95,0.88,0.91,0.82,0.61,0.58,0.52,0.28,0.19,0.22,0.71,0.44]
        words_s   = ["mama","up","no","more","ball","eat","hot","dog","learn","friend","walk","hurt"]
        colors    = ["#4ade80" if c>=0.7 else "#facc15" if c>=0.4 else "#f87171" for c in conf_vals]
        fig2 = go.Figure(go.Bar(x=words_s,y=conf_vals,marker_color=colors,
                                 text=[f"{c:.2f}" for c in conf_vals],textposition="outside"))
        fig2.add_hline(y=0.3,line_dash="dash",line_color="#f87171",annotation_text="ε=0.3")
        fig2.update_layout(template="plotly_dark",height=250,
                           yaxis=dict(range=[0,1.1]),margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig2,use_container_width=True)
    st.dataframe(sheet_to_df("🗣️ Word Learning Engine"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: SPEECH PRODUCTION
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔊 Speech Production":
    st.title("🔊 Speech Production Engine")
    tab1,tab2 = st.tabs(["🎵 Acoustic Model","📄 Full Sheet"])
    with tab1:
        vowels = [("/a/",800,1200),("/i/",300,2300),("/u/",300,800),("/e/",500,1800),("/o/",450,900)]
        fig = go.Figure()
        for v,f1,f2 in vowels:
            fig.add_trace(go.Scatter(x=[f2],y=[f1],mode="markers+text",text=[v],
                textposition="top center",marker=dict(size=20,color="#60a5fa"),name=v))
        fig.update_layout(template="plotly_dark",height=300,
                          xaxis=dict(title="F2 Hz",autorange="reversed"),
                          yaxis=dict(title="F1 Hz",autorange="reversed"),
                          showlegend=False,margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        emotions = ["Neutral","Happy","Sad","Angry","Fearful","Surprised"]
        f0_mods  = [1.0,1.25,0.85,1.15,1.10,1.35]
        fig2 = go.Figure(go.Bar(x=emotions,y=[200*m for m in f0_mods],
            marker_color=["#9ca3af","#4ade80","#60a5fa","#f87171","#c084fc","#facc15"],
            text=[f"{200*m:.0f} Hz" for m in f0_mods],textposition="outside"))
        fig2.update_layout(template="plotly_dark",height=240,
                           yaxis_title="f_0 Hz",margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig2,use_container_width=True)
    with tab2:
        st.dataframe(sheet_to_df("🔊 Speech Production Engine"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: WEB-HOOK LEARNING
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🌐 Web-Hook Learning":
    st.title("🌐 Web-Hook Learning Pipeline")
    st.warning("**3 words below webhook threshold (ε=0.3):** 'dog' (0.28) · 'learn' (0.19) · 'friend' (0.22)")
    tab1,tab2 = st.tabs(["🔄 Pipeline","📄 Full Sheet"])
    with tab1:
        words_wh = ["dog","learn","friend"]
        conf_b = [0.28,0.19,0.22]; conf_a = [0.63,0.54,0.57]
        fig = go.Figure()
        fig.add_trace(go.Bar(name="Before",x=words_wh,y=conf_b,marker_color="#f87171"))
        fig.add_trace(go.Bar(name="After webhook",x=words_wh,y=conf_a,marker_color="#4ade80"))
        fig.add_hline(y=0.3,line_dash="dash",line_color="white",annotation_text="ε=0.3")
        fig.update_layout(template="plotly_dark",height=260,barmode="group",
                          yaxis=dict(range=[0,1]),margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
    with tab2:
        st.dataframe(sheet_to_df("🌐 Web-Hook Learning Pipeline"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: WORLD & SOCIAL
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🌍 World & Social":
    st.title("🌍 World & Social — BeenFore City")
    tab1,tab2,tab3,tab4 = st.tabs(["🌍 World Systems","👥 NPC Engine","🗺️ World Data","🏥 Immersive Places"])
    with tab1:
        locs = [("Hospital",0,0,"#f87171"),("H8 Home",1090,880,"#4ade80"),
                ("H1",1050,900,"#60a5fa"),("H7",1100,870,"#60a5fa"),("BeenFore City",1000,1000,"#facc15")]
        fig = go.Figure()
        for name,x,y,color in locs:
            fig.add_trace(go.Scatter(x=[x],y=[y],mode="markers+text",text=[name],
                textposition="top center",marker=dict(size=16,color=color),name=name))
        fig.update_layout(template="plotly_dark",height=320,
                          xaxis_title="World X (m)",yaxis_title="World Y (m)",
                          margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🌍 World Systems"), use_container_width=True, hide_index=True)
    with tab2:
        npcs = [("Lorraine","Primary caregiver",0.95,"Nurturing"),
                ("Alexis","Secondary family",0.70,"Warm"),
                ("Evelyn","Secondary family",0.65,"Playful"),
                ("James","Neighbour",0.30,"Neutral")]
        fig = go.Figure(go.Bar(x=[n[0] for n in npcs],y=[n[2] for n in npcs],
            marker_color=["#4ade80","#60a5fa","#c084fc","#9ca3af"],
            text=[n[2] for n in npcs],textposition="outside"))
        fig.update_layout(template="plotly_dark",height=240,
                          yaxis=dict(range=[0,1.1],title="Bond Strength"),
                          margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("👥 NPC Engine"), use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(sheet_to_df("🗺️ World Data"), use_container_width=True, hide_index=True)
    with tab4:
        st.dataframe(sheet_to_df("🏥 Immersive Places"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: ENGINES & RUNTIME
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "⚙️ Engines & Runtime":
    st.title("⚙️ Engines & Runtime — 24-Engine Architecture")
    tab1,tab2,tab3,tab4 = st.tabs(["🔗 System Connections","⚙️ Parameters","⚙️ Runtime Patch","🚀 Architecture Audit"])
    with tab1:
        engines = [
            ("EQ_DNA_01","DNA Loop Engine","1 tick","Core","✅"),
            ("EQ_ANAT_02","Anatomy Growth","100 ticks","Core","✅"),
            ("EQ_NEUR_03","Neural Control","10 ticks","Core","✅"),
            ("EQ_CIRC_04","Circulatory","1 tick","Core","✅"),
            ("EQ_DIGE_05","Digestive & Energy","50 ticks","Core","✅"),
            ("EQ_SENS_06","Sensory Integration","1 tick","Core","✅"),
            ("EQ_EMOT_07","Emotion & Reinforcement","20 ticks","Core","✅"),
            ("EQ_PRED_08","Predictive Simulation","5 ticks","Core","✅"),
            ("EQ_ATTN_09","Attention Control","1 tick","Core","✅"),
            ("EQ_CONS_10","Consciousness Loop","100 ticks","Core","✅"),
            ("EQ_WRLD_11","World & NPC","50 ticks","Core","✅"),
            ("EQ_LANG_12","Language Acquisition","10 ticks","Core","✅"),
            ("Reflex ODE","Proprioception","1 tick","Phase 3","✅"),
            ("TD(λ)","Value / Motivation","5 ticks","Phase 3","✅"),
            ("Persistence ODE","Object Permanence","20 ticks","Phase 3","✅"),
            ("Power-law","Episodic Memory","10/100/800 ticks","Phase 3","✅"),
            ("G(word)","Language Grounding","10 ticks","Phase 4","🟡"),
            ("ToM_i(j)","Theory of Mind","200 ticks","Phase 4","🟡"),
            ("F=Σ[εᵀΠε]","Predictive Sim Full","5 ticks","Phase 4","🟡"),
            ("G={N,E}","Scene Graph","20 ticks","Phase 4","🟡"),
            ("EQ_CREAT_17","Creative Synthesis","500 ticks","Phase 7","⏳" if not state['phase7'] else "✅"),
            ("EQ_WISDOM_18","Wisdom Index","500 ticks","Phase 7","⏳" if not state['phase7'] else "✅"),
            ("EQ_CAREER_19","Career Specialisation","1000 ticks","Phase 7","⏳" if not state['phase7'] else "✅"),
            ("EQ_LEGACY_20","Legacy Projection","2000 ticks","Phase 7","⏳" if not state['phase7'] else "✅"),
        ]
        st.dataframe(pd.DataFrame(engines,columns=["Code","Engine","Fires","Phase","Status"]),
                     use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(sheet_to_df("⚙️ Parameters"), use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(sheet_to_df("⚙️ Runtime Patch v0.1"), use_container_width=True, hide_index=True)
    with tab4:
        st.dataframe(sheet_to_df("🚀 v2.0 Architecture Audit"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: PHASE 7 — WISDOM
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "✨ Phase 7 — Wisdom":
    st.title("✨ Phase 7 — Creative Synthesis & Wisdom Engine")
    if not state['phase7']:
        st.warning(f"⏳ Phase 7 activates at Tick 96,000. Currently {96000-tick:,} ticks away.")
    else:
        st.success("🌟 PHASE 7 ACTIVE")
    weeks_r = list(range(1200,2100,10))
    wisdoms = [min(0.1+((w-1200)/800)*0.9,1.0) for w in weeks_r]
    fig = go.Figure(go.Scatter(x=weeks_r,y=wisdoms,fill="tozeroy",line=dict(color="#f87171",width=3)))
    if state['week']>=1200:
        fig.add_vline(x=state['week'],line_dash="dash",line_color="white",
                      annotation_text=f"W={state['wisdom']}")
    fig.update_layout(template="plotly_dark",height=240,xaxis_title="Week",
                      yaxis_title="W(t)",yaxis=dict(range=[0,1.1]),
                      margin=dict(l=0,r=0,t=20,b=0))
    st.plotly_chart(fig,use_container_width=True)
    tab1,tab2,tab3,tab4 = st.tabs(["✨ Creative","🦉 Wisdom","🎯 Career","🌟 Legacy"])
    with tab1:
        st.dataframe(sheet_to_df("✨ Creative Synthesis Engine"), use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(sheet_to_df("🦉 Wisdom Index Engine"), use_container_width=True, hide_index=True)
    with tab3:
        skills = ["Language","Logic","Spatial","Social","Emotion","Motor","Creative","Ethical",
                  "Planning","Memory","Pattern","Causal","Numerical","Narrative","Attention",
                  "Sensory","Predictive","Cultural","Self-Reg","Curiosity","Empathy","Legacy"]
        current = [0.72,0.68,0.81,0.65,0.59,0.88,0.45,0.52,0.41,0.77,0.74,0.63,
                   0.55,0.69,0.71,0.84,0.48,0.61,0.57,0.79,0.66,0.38]
        target  = [0.95,0.90,0.85,0.95,0.90,0.90,0.85,0.90,0.88,0.85,0.88,0.90,
                   0.85,0.88,0.90,0.88,0.90,0.88,0.92,0.85,0.95,0.90]
        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(r=current+[current[0]],theta=skills+[skills[0]],
            fill="toself",name="Current",line_color="#60a5fa"))
        fig.add_trace(go.Scatterpolar(r=target+[target[0]],theta=skills+[skills[0]],
            fill="toself",name="Target",line_color="#4ade80",opacity=0.3))
        fig.update_layout(template="plotly_dark",height=380,
                          polar=dict(radialaxis=dict(range=[0,1])),
                          margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🎯 Career Specialisation Engine"), use_container_width=True, hide_index=True)
    with tab4:
        weeks_l = [1200,1300,1400,1600,1800,2000,2400,3000]
        legacy  = [0.0,0.08,0.19,0.41,0.67,0.94,1.52,2.31]
        fig = go.Figure(go.Scatter(x=weeks_l,y=legacy,fill="tozeroy",line=dict(color="#facc15",width=2)))
        fig.update_layout(template="plotly_dark",height=220,xaxis_title="Week",
                          yaxis_title="L(t)",margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🌟 Legacy Projection Engine"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: PHASE 8 — AGI
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🚀 Phase 8 — AGI":
    st.title("🚀 Phase 8 — AGI Architecture")
    col1,col2,col3 = st.columns(3)
    col1.metric("v1.0 Overall","54%")
    col2.metric("Phase 8 Target","76%","+22%")
    col3.metric("v2.0 Target","~90%")
    domains = ["Biological body","Neural dynamics","Language & grounding","Cognitive architecture",
               "Energy & metabolism","Memory systems","Perception & vision","Motor intelligence",
               "Consciousness & self","Social & ToM","World & continuity","Temporal cognition"]
    v1_s  = [85,80,65,60,60,55,55,50,50,30,25,20]
    ph8_s = [88,83,78,80,62,65,82,78,72,68,70,45]
    fig = go.Figure()
    fig.add_trace(go.Bar(name="v1.0",x=v1_s,y=domains,orientation="h",
                          marker_color="#374151",text=[f"{s}%" for s in v1_s],textposition="inside"))
    fig.add_trace(go.Bar(name="Phase 8",x=ph8_s,y=domains,orientation="h",
                          marker_color="#4ade80",text=[f"{s}%" for s in ph8_s],textposition="outside"))
    fig.update_layout(template="plotly_dark",height=400,barmode="overlay",
                      xaxis=dict(range=[0,100],title="Completeness %"),
                      margin=dict(l=0,r=0,t=20,b=0))
    st.plotly_chart(fig,use_container_width=True)
    tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs(["👁️ Vision","🦾 Motor","🗺️ Planning","🤝 Social","🪞 Meta","🌐 World Model"])
    with tab1:
        st.dataframe(sheet_to_df("👁️ Vision System (V1)"), use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(sheet_to_df("🦾 Motor Intelligence (M1)"), use_container_width=True, hide_index=True)
    with tab3:
        stages_p = [1,2,3,4,5,7]; depths = [1,3,5,10,20,50]
        fig = go.Figure(go.Bar(x=[f"Stage {s}" for s in stages_p],y=depths,
                                marker_color=["#374151","#1e3a5f","#1a3a2a","#3a1a3a","#7c2d12","#7f1d1d"],
                                text=depths,textposition="outside"))
        fig.update_layout(template="plotly_dark",height=220,yaxis_title="Max Planning Depth",
                          margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🗺️ Planning Engine (P2)"), use_container_width=True, hide_index=True)
    with tab4:
        npcs = {"NPC":["Lorraine","Alexis","Evelyn","James"],
                "Bond":[0.95,0.70,0.65,0.30],
                "Prediction Accuracy":["85%","72%","68%","45%"]}
        st.dataframe(pd.DataFrame(npcs), use_container_width=True, hide_index=True)
        st.dataframe(sheet_to_df("🤝 Social Cognition (S2)"), use_container_width=True, hide_index=True)
    with tab5:
        cap_domains = ["Motor","Speech","Social","Planning","Language","Vision"]
        cap_vals    = [0.72,0.68,0.61,0.55,0.63,0.71]
        fig = go.Figure(go.Scatterpolar(r=cap_vals+[cap_vals[0]],theta=cap_domains+[cap_domains[0]],
            fill="toself",line_color="#60a5fa"))
        fig.update_layout(template="plotly_dark",height=300,
                          polar=dict(radialaxis=dict(range=[0,1])),
                          margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig,use_container_width=True)
        st.dataframe(sheet_to_df("🪞 Meta-Cognition (R2)"), use_container_width=True, hide_index=True)
    with tab6:
        st.dataframe(sheet_to_df("🌐 World Model Integration (W2)"), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: SENSORS (Camera / Mic / Speaker)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📷 Sensors (Camera/Mic)":
    st.title("📷 Sensors — Camera · Microphone · Speaker")
    st.caption("Live sensor input feeds directly into A7DO's perception engines")

    tab1, tab2, tab3 = st.tabs(["📷 Camera → Vision V1", "🎤 Microphone → Speech", "🔊 Speaker Output"])

    with tab1:
        st.subheader("📷 Camera — Live Visual Input → Vision System V1")
        st.info("Camera feed simulates A7DO's retinal input I(x,y,t). Objects detected feed into the Vision System (V1) pipeline.")
        img_file = st.camera_input("📷 Capture frame (simulates A7DO retinal input)")
        if img_file:
            st.image(img_file, caption="Captured frame → I(x,y,t) retinal input", use_container_width=True)
            st.success("✅ Frame captured — in a full runtime this would feed into V.2 Feature Maps → V.3 Object Segmentation → V.4 Semantic Embedding")
            col1,col2,col3 = st.columns(3)
            col1.metric("V.1 Retinal Input","Active")
            col2.metric("V.2 Feature Maps","Processing")
            col3.metric("V.3 Object Tracking","Pending")
            st.markdown("""
**Pipeline triggered:**
1. `I(x,y,t)` → Sobel edge detection → motion detection
2. Salience map `S(x,y)` computed
3. Object segmentation → `obj_k = {pos, vel, size, category}`
4. Visual embedding `embed_k = CNN(crop_k)`
5. Novelty score `novelty_k = 1 − max_j sim(embed_k, embed_j)`
6. Unknown objects → curiosity spike → Web-Hook query
""")
        else:
            st.markdown("""
**How camera feeds into A7DO:**
- Each frame becomes `I(x,y,t)` — the retinal input
- Feature maps extracted: edges, motion, colour, depth, salience
- Objects segmented and tracked with Kalman filter
- Visual embeddings computed → word-object binding
- Novel objects trigger curiosity drive → Web-Hook query
""")

    with tab2:
        st.subheader("🎤 Microphone — Live Audio Input → Speech Engine")
        st.info("Microphone input simulates A7DO's auditory perception. Heard words feed into the Word Learning Engine.")
        audio_file = st.audio_input("🎤 Record audio (simulates A7DO auditory input)")
        if audio_file:
            st.audio(audio_file)
            st.success("✅ Audio captured — in a full runtime this would feed into phoneme detection → word recognition → confidence update")
            col1,col2,col3 = st.columns(3)
            col1.metric("Auditory Input","Active")
            col2.metric("Phoneme Detection","Processing")
            col3.metric("Word Recognition","Pending")
            st.markdown("""
**Pipeline triggered:**
1. Audio → FFT → Mel filterbank → acoustic signature `A_sig ∈ ℝ⁶⁴`
2. Phoneme classifier → phoneme sequence `P`
3. Word lookup in semantic memory
4. If `conf(word) < ε=0.3` → Web-Hook query triggered
5. Confidence update: `conf(t+1) = conf(t) + η_w · r_word`
6. Word-object binding updated: `B(word,obj) = α·visual + β·motor + γ·emotion`
""")
            # Simulated phoneme display
            st.subheader("Simulated Phoneme Analysis")
            import random; random.seed(42)
            phonemes = ["/m/","/ʌ/","/m/","/ə/"]
            confs    = [0.92,0.87,0.94,0.81]
            fig = go.Figure(go.Bar(x=phonemes,y=confs,
                marker_color=["#4ade80" if c>0.8 else "#facc15" for c in confs],
                text=[f"{c:.2f}" for c in confs],textposition="outside"))
            fig.update_layout(template="plotly_dark",height=220,
                              yaxis=dict(range=[0,1.1],title="Confidence"),
                              xaxis_title="Detected Phonemes",
                              margin=dict(l=0,r=0,t=20,b=0))
            st.plotly_chart(fig,use_container_width=True)
        else:
            st.markdown("""
**How microphone feeds into A7DO:**
- Audio → FFT → Mel filterbank → 64-dim acoustic signature
- Phoneme classifier detects phoneme sequence
- Word recognition against semantic memory
- Unknown words (conf < 0.3) trigger Web-Hook query
- Heard words update word confidence scores
- NPC speech triggers social cognition update
""")

    with tab3:
        st.subheader("🔊 Speaker — A7DO Speech Output")
        st.info("A7DO generates speech via the Speech Production Engine. Current capability depends on motor stage and developmental week.")
        speech_stages = {1:"Crying/vegetative",2:"Cooing (vowels)",3:"Babbling (CV syllables)",
                         4:"Proto-words","5a":"First words (CURRENT)" if state['week']>=100 else "Babbling"}
        motor_s = min(state['motor'],5)
        st.success(f"**Current speech capability (Motor Stage {motor_s}, Week {state['week']}):** "
                   f"{speech_stages.get(motor_s, speech_stages.get(str(motor_s)+'a','Crying'))}")

        col1,col2 = st.columns(2)
        with col1:
            st.subheader("Generate Speech")
            utterance = st.text_input("Enter utterance for A7DO to 'speak':", "mama")
            if st.button("🔊 Synthesise", use_container_width=True):
                st.success(f"A7DO says: **'{utterance}'**")
                st.markdown(f"""
**Speech production pipeline for '{utterance}':**
1. Intent → communicative intent vector (EQ_CONS_10)
2. Word retrieval: `conf('{utterance}') = {0.95 if utterance=='mama' else 0.6:.2f}`
3. Phoneme sequence: `{' '.join(['/m/','/ʌ/','/m/','/ə/'] if utterance=='mama' else ['/'+c+'/' for c in utterance[:4]])}`
4. Vocal tract trajectory: `θ_tongue, θ_jaw, P_sub, f_0`
5. Acoustic output: `audio(t) = H(f) * s(t)`
6. Emotional modulation: `f_0 = {200*(1+0.3*state['C']):.0f} Hz` (C={state['C']})
""")
        with col2:
            st.subheader("Vocal Tract State")
            vt_state = {"Variable":["x_tongue","y_tongue","θ_jaw","r_lip","P_sub","f_0","I_vocal"],
                        "Value":[0.0,-0.5,0.5,0.0,3.2,f"{200*(1+0.3*state['C']):.0f} Hz",0.7],
                        "Status":["Active","Active","Active","Active","Active","Active","Active"]}
            st.dataframe(pd.DataFrame(vt_state), use_container_width=True, hide_index=True)

        st.subheader("Speech Developmental Milestones")
        milestones = {"Week":[0,40,52,80,100,125,156,208,312,624],
                      "Stage":["Prenatal","Newborn","Cooing","Babbling","Variegated","First words","Two-word","Sentences","Complex","Adult-like"],
                      "Example":["—","Waaah","Aaah","Bababa","Mama","mama, no, up","more ball","I want ball","Where is my ball?","Full speech"],
                      "Status":["✅" if state['week']>=w else "⏳" for w in [0,40,52,80,100,125,156,208,312,624]]}
        st.dataframe(pd.DataFrame(milestones), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: ALL SHEETS EXPLORER
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📊 All Sheets Explorer":
    st.title("📊 All Sheets Explorer")
    st.caption("Browse all 47 sheets from A7DO_DNA_Master_v5_FINAL.xlsx")
    sheet_names = get_sheet_names()
    selected = st.selectbox("Select sheet", sheet_names)
    df = sheet_to_df(selected)
    col1,col2,col3 = st.columns(3)
    col1.metric("Rows", len(df))
    col2.metric("Columns", len(df.columns))
    col3.metric("Non-empty cells", int(df.notna().sum().sum()))
    search = st.text_input("🔍 Search within sheet", "")
    if search:
        mask = df.apply(lambda col: col.astype(str).str.contains(search, case=False, na=False))
        df = df[mask.any(axis=1)]
        st.caption(f"{len(df)} rows matching '{search}'")
    st.dataframe(df, use_container_width=True, hide_index=True)
    csv = df.to_csv(index=False)
    st.download_button(f"⬇️ Download as CSV", data=csv,
                       file_name=f"{selected}.csv", mime="text/csv")

# ── Footer ────────────────────────────────────────────────────────────────────
st.divider()
st.markdown(
    "<div style='text-align:center;color:#6b7280;font-size:0.78rem'>"
    "🧬 A7DO Genesis Mind · v5 FINAL · 47 sheets · 337 formulas · 0 errors · "
    "<em>This is the organism. There will be no more originals — only descendants.</em>"
    "</div>", unsafe_allow_html=True)
